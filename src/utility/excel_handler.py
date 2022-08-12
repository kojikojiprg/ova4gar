import os

import openpyxl
import pandas as pd


def save(
    excel_path: str,
    sheet_name: str,
    df: pd.DataFrame,
    index: bool = False,
    header: bool = False,
):
    # write dataframe for excel
    if os.path.exists(excel_path):
        # delete sheet if same sheet exists
        workbook = openpyxl.load_workbook(filename=excel_path)
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
            if len(workbook.sheetnames) == 0:
                workbook.create_sheet("dmy")  # avoid error
            workbook.save(excel_path)
        workbook.close()

        # over write excel file
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name, index=index, header=header)

        workbook = openpyxl.load_workbook(filename=excel_path)
        if "dmy" in workbook.sheetnames:
            workbook.remove(workbook["dmy"])
            workbook.save(excel_path)
        workbook.close()
    else:
        # create and write excel file
        df.to_excel(excel_path, sheet_name, index=index, header=header)
